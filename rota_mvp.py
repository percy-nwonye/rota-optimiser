# rota_mvp.py
# FULL REPLACEMENT FILE (Config + CSV driven)
#
# Features:
# - Loads config.json (auto-creates with defaults)
# - Loads staff.csv + requirements.csv if present; otherwise uses demo dataset
# - Hard blocks:
#   - cannot assign same staff to same shift on same date (NON-OVERRIDABLE)
#   - cannot assign nights to staff who can't do nights (NON-OVERRIDABLE)
# - Soft warnings:
#   - one_shift_per_day -> "Already working that date (double shift)"
#   - consecutive limits -> "Would exceed consecutive ..."
# - Override Mode:
#   - interactive selection for unfilled slots
#   - policy gates controlled by config.json
#   - bank_only_consecutive_override option supported
# - Exports:
#   - rota_assignments.csv, rota_unfilled.csv, rota_fairness.csv
#   - rota_overrides.csv, rota_explainability.csv, rota_audit.csv
#   - rota.xlsx (Assignments/Unfilled/Fairness/Overrides/Explainability/Audit) if openpyxl available
# - Explainability v2: top-N ranking + blocks per assignment decision
# - Quality Gate: PASS/FAIL based on config policy (unfilled + policy violations)

from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Set, Tuple, Any

DATE_FMT = "%Y-%m-%d"
TIME_FMT = "%H:%M"


# =========================
# DATA MODELS
# =========================
@dataclass(frozen=True)
class ShiftTemplate:
    id: str
    name: str
    start: str
    end: str
    is_night: bool


@dataclass(frozen=True)
class Staff:
    id: str
    name: str
    role: str
    contract_type: str   # "permanent" or "bank"
    target_hours_per_week: int
    can_do_nights: bool


@dataclass(frozen=True)
class Requirement:
    date: str
    shift_id: str
    role: str
    required: int


@dataclass(frozen=True)
class Assignment:
    date: str
    shift_id: str
    role: str
    staff_id: str


@dataclass(frozen=True)
class UnfilledSlot:
    date: str
    shift_id: str
    role: str
    reason: str


@dataclass(frozen=True)
class CoverCandidate:
    staff_id: str
    score: float
    blocked_by: List[str]


@dataclass(frozen=True)
class OverrideEvent:
    timestamp_utc: str
    date: str
    shift_id: str
    shift_name: str
    role: str
    staff_id: str
    staff_name: str
    contract_type: str
    is_override: bool
    is_manual_cover: bool
    override_double_shift: bool
    override_consecutive: bool
    override_reason_text: str


@dataclass(frozen=True)
class ExplainabilityEvent:
    timestamp_utc: str
    date: str
    shift_id: str
    shift_name: str
    role: str
    chosen_staff_id: str
    chosen_staff_name: str
    chosen_score: float
    decision_type: str  # "auto" or "manual"
    top_n: int
    ranked_candidates_json: str
    notes: str


@dataclass(frozen=True)
class AuditRow:
    staff_id: str
    staff_name: str
    role: str
    contract_type: str
    target_hours_per_week: int
    hours: float
    nights: int
    weekends: int
    shifts_total: int
    days_worked: int
    double_shift_days: int
    max_consecutive_nights_actual: int
    max_consecutive_long_days_actual: int
    overload_flags: str


# =========================
# CONFIG LOADING
# =========================
def load_config(path: str = "config.json") -> Dict:
    # Defaults are permissive enough to cover the rota in override mode,
    # but still fully policy-controlled via config.json.
    defaults = {
        "constraints": {
            "one_shift_per_day": True,
            "max_consecutive_nights": 2,
            "max_consecutive_long_days": 3
        },
        "overrides": {
            "allow_overrides": True,
            "interactive_overrides": True,
            "warnings_enabled": True,

            "require_warning_ack": True,
            "warning_ack_phrase": "I UNDERSTAND",

            # policy switches (enforced by code)
            "allow_double_shift": True,
            "allow_exceed_consecutive_nights": True,
            "allow_exceed_consecutive_long_days": True,

            # if true: consecutive override suggestions/permission only for BANK staff
            "bank_only_consecutive_override": False,
        },
        "scoring_weights": {
            "hours": 1.0,
            "nights": 15.0,
            "weekends": 10.0
        },
        "exports": {
            "export_folder": ".",
            "export_csv": True,
            "export_excel": True,
            "export_overrides_csv": True,
            "export_explainability_csv": True,
            "export_audit_csv": True,
        },
        "cover_suggestions": {
            "top_n": 3
        },
        "explainability": {
            "top_n_per_decision": 5
        },
        "debug": {
            "print_no_candidate_ranking": False
        }
    }

    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(defaults, f, indent=2)
        print(f"⚠️  {path} not found, created a default one. You can edit it anytime.")
        return defaults

    with open(path, "r", encoding="utf-8") as f:
        user_cfg = json.load(f)

    def merge(d: Dict, u: Dict) -> Dict:
        out = dict(d)
        for k, v in u.items():
            if isinstance(v, dict) and isinstance(out.get(k), dict):
                out[k] = merge(out[k], v)
            else:
                out[k] = v
        return out

    merged = merge(defaults, user_cfg)

    # ensure nested keys always exist
    merged.setdefault("explainability", {})
    merged["explainability"].setdefault("top_n_per_decision", defaults["explainability"]["top_n_per_decision"])
    merged.setdefault("debug", {})
    merged["debug"].setdefault("print_no_candidate_ranking", defaults["debug"]["print_no_candidate_ranking"])
    merged.setdefault("overrides", {})
    merged["overrides"].setdefault("bank_only_consecutive_override", defaults["overrides"]["bank_only_consecutive_override"])

    return merged


# =========================
# CLI HELPERS
# =========================
def ask_yes_no(prompt: str, default_no: bool = True) -> bool:
    suffix = " [y/N]: " if default_no else " [Y/n]: "
    while True:
        ans = input(prompt + suffix).strip().lower()
        if ans == "":
            return (not default_no)
        if ans in ("y", "yes"):
            return True
        if ans in ("n", "no"):
            return False
        print("Please type y or n.")


def ask_choice(prompt: str, choices: List[str], allow_skip: bool = True) -> Optional[int]:
    print(prompt)
    for i, c in enumerate(choices, 1):
        print(f"  {i}) {c}")
    if allow_skip:
        print("  0) Skip")

    while True:
        raw = input("Select: ").strip()
        if raw == "" and allow_skip:
            return None
        if raw.isdigit():
            n = int(raw)
            if allow_skip and n == 0:
                return None
            if 1 <= n <= len(choices):
                return n - 1
        print("Enter a valid number.")


# =========================
# CSV LOADERS
# =========================
def _read_bool(x: str) -> bool:
    return str(x).strip().lower() in ("1", "true", "yes", "y", "t")


def load_staff_csv(path: str = "staff.csv") -> List[Staff]:
    staff: List[Staff] = []
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        required_cols = {"id", "name", "role", "contract_type", "target_hours_per_week", "can_do_nights"}
        missing = required_cols - set(r.fieldnames or [])
        if missing:
            raise ValueError(f"staff.csv missing columns: {sorted(missing)}")

        for row in r:
            staff.append(
                Staff(
                    id=row["id"].strip(),
                    name=row["name"].strip(),
                    role=row["role"].strip(),
                    contract_type=row["contract_type"].strip().lower(),
                    target_hours_per_week=int(row["target_hours_per_week"]),
                    can_do_nights=_read_bool(row["can_do_nights"]),
                )
            )
    return staff


def load_requirements_csv(path: str = "requirements.csv") -> List[Requirement]:
    reqs: List[Requirement] = []
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        required_cols = {"date", "shift_id", "role", "required"}
        missing = required_cols - set(r.fieldnames or [])
        if missing:
            raise ValueError(f"requirements.csv missing columns: {sorted(missing)}")

        for row in r:
            reqs.append(
                Requirement(
                    date=row["date"].strip(),
                    shift_id=row["shift_id"].strip(),
                    role=row["role"].strip(),
                    required=int(row["required"]),
                )
            )
    return reqs


# =========================
# CORE ENGINE
# =========================
class RotaGenerator:
    def __init__(
        self,
        config: Dict,
        shift_templates: Dict[str, ShiftTemplate],
        staff: List[Staff],
        requirements: List[Requirement],
    ):
        self.cfg = config
        self.shift_templates = shift_templates
        self.staff: Dict[str, Staff] = {s.id: s for s in staff}
        self.requirements = requirements

        self.assignments: List[Assignment] = []
        self.unfilled: List[UnfilledSlot] = []
        self.override_events: List[OverrideEvent] = []
        self.explainability_events: List[ExplainabilityEvent] = []

        self.staff_hours: Dict[str, float] = {s.id: 0.0 for s in staff}
        self.staff_nights: Dict[str, int] = {s.id: 0 for s in staff}
        self.staff_weekends: Dict[str, int] = {s.id: 0 for s in staff}

        self.staff_working_on_date: Dict[str, Set[str]] = {}  # date -> set(staff_id)
        self._assignments_by_date: Dict[str, List[Assignment]] = {}  # date -> assignments

        # streak tracking (used during auto generation; override streak checks use projected scan)
        self.consec_nights: Dict[str, int] = {s.id: 0 for s in staff}
        self.consec_long: Dict[str, int] = {s.id: 0 for s in staff}
        self.last_date: Dict[str, Optional[str]] = {s.id: None for s in staff}
        self.last_shift_was_night: Dict[str, bool] = {s.id: False for s in staff}
        self.last_shift_was_long: Dict[str, bool] = {s.id: False for s in staff}

    # ---------- helpers ----------
    @staticmethod
    def _now_utc_iso() -> str:
        return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    @staticmethod
    def _is_weekend(date_str: str) -> bool:
        return datetime.strptime(date_str, DATE_FMT).weekday() >= 5

    @staticmethod
    def _is_next_day(prev: str, curr: str) -> bool:
        d1 = datetime.strptime(prev, DATE_FMT)
        d2 = datetime.strptime(curr, DATE_FMT)
        return (d2 - d1).days == 1

    @staticmethod
    def _shift_hours(shift: ShiftTemplate) -> float:
        start = datetime.strptime(shift.start, TIME_FMT)
        end = datetime.strptime(shift.end, TIME_FMT)
        if end <= start:
            end = end + timedelta(days=1)
        return (end - start).seconds / 3600.0

    def _score(self, staff_id: str) -> float:
        w = self.cfg["scoring_weights"]
        return (
            self.staff_hours[staff_id] * float(w.get("hours", 1.0))
            + self.staff_nights[staff_id] * float(w.get("nights", 15.0))
            + self.staff_weekends[staff_id] * float(w.get("weekends", 10.0))
        )

    def _debug_enabled(self) -> bool:
        return bool(self.cfg.get("debug", {}).get("print_no_candidate_ranking", False))

    def _bank_only_consecutive_override_enabled(self) -> bool:
        return bool(self.cfg.get("overrides", {}).get("bank_only_consecutive_override", False))

    # ---------- logs ----------
    def _log_override(
        self,
        *,
        date: str,
        shift: ShiftTemplate,
        role: str,
        staff_id: str,
        is_override: bool,
        is_manual_cover: bool,
        override_double: bool,
        override_consecutive: bool,
        reason_text: str,
    ) -> None:
        s = self.staff[staff_id]
        self.override_events.append(
            OverrideEvent(
                timestamp_utc=self._now_utc_iso(),
                date=date,
                shift_id=shift.id,
                shift_name=shift.name,
                role=role,
                staff_id=staff_id,
                staff_name=s.name,
                contract_type=s.contract_type,
                is_override=is_override,
                is_manual_cover=is_manual_cover,
                override_double_shift=override_double,
                override_consecutive=override_consecutive,
                override_reason_text=reason_text,
            )
        )

    def _log_explainability(
        self,
        *,
        date: str,
        shift: ShiftTemplate,
        role: str,
        chosen_staff_id: str,
        decision_type: str,
        ranked_candidates: List[Dict[str, Any]],
        notes: str,
    ) -> None:
        chosen = self.staff[chosen_staff_id]
        top_n = int(self.cfg.get("explainability", {}).get("top_n_per_decision", 5))
        self.explainability_events.append(
            ExplainabilityEvent(
                timestamp_utc=self._now_utc_iso(),
                date=date,
                shift_id=shift.id,
                shift_name=shift.name,
                role=role,
                chosen_staff_id=chosen_staff_id,
                chosen_staff_name=chosen.name,
                chosen_score=float(self._score(chosen_staff_id)),
                decision_type=decision_type,
                top_n=top_n,
                ranked_candidates_json=json.dumps(ranked_candidates, ensure_ascii=False),
                notes=notes,
            )
        )

    # ---------- constraint helpers ----------
    def _staff_worked_shift_type_on_date(self, sid: str, date: str, *, want_night: bool) -> bool:
        for a in self._assignments_by_date.get(date, []):
            if a.staff_id != sid:
                continue
            sh = self.shift_templates[a.shift_id]
            if sh.is_night == want_night:
                return True
        return False

    def _projected_night_streak(self, sid: str, date: str) -> int:
        streak = 1
        d = datetime.strptime(date, DATE_FMT)
        while True:
            prev = (d - timedelta(days=1)).strftime(DATE_FMT)
            if self._staff_worked_shift_type_on_date(sid, prev, want_night=True):
                streak += 1
                d = d - timedelta(days=1)
                continue
            break
        return streak

    def _projected_long_streak(self, sid: str, date: str) -> int:
        streak = 1
        d = datetime.strptime(date, DATE_FMT)
        while True:
            prev = (d - timedelta(days=1)).strftime(DATE_FMT)
            if self._staff_worked_shift_type_on_date(sid, prev, want_night=False):
                streak += 1
                d = d - timedelta(days=1)
                continue
            break
        return streak

    def _blocked_reasons(
        self,
        sid: str,
        role: str,
        shift: ShiftTemplate,
        date: str,
        *,
        ignore_one_shift_per_day: bool = False,
        ignore_consecutive_rules: bool = False,
    ) -> List[str]:
        c = self.cfg["constraints"]
        reasons: List[str] = []
        s = self.staff[sid]

        if s.role != role:
            reasons.append("Role mismatch")

        # HARD: same staff cannot be assigned twice to SAME shift on SAME date
        for a in self._assignments_by_date.get(date, []):
            if a.staff_id == sid and a.shift_id == shift.id:
                reasons.append("Already assigned to this shift (NON-OVERRIDABLE)")
                break

        # HARD: night capability
        if shift.is_night and not s.can_do_nights:
            reasons.append("Cannot do nights (NON-OVERRIDABLE)")

        # one shift per day (soft)
        if c.get("one_shift_per_day", True) and not ignore_one_shift_per_day:
            if sid in self.staff_working_on_date.get(date, set()):
                reasons.append("Already working that date (double shift)")

        # consecutive limits (soft)
        if not ignore_consecutive_rules:
            if shift.is_night:
                proj = self._projected_night_streak(sid, date)
                if proj > int(c["max_consecutive_nights"]):
                    reasons.append(f"Would exceed consecutive nights ({proj} > {c['max_consecutive_nights']})")
            else:
                proj = self._projected_long_streak(sid, date)
                if proj > int(c["max_consecutive_long_days"]):
                    reasons.append(f"Would exceed consecutive long days ({proj} > {c['max_consecutive_long_days']})")

        return reasons

    def _rank_candidates_for_decision(self, role: str, shift: ShiftTemplate, date: str, top_n: int) -> List[Dict[str, Any]]:
        ranked: List[Dict[str, Any]] = []
        for sid, s in self.staff.items():
            if s.role != role:
                continue
            blocked = self._blocked_reasons(sid, role, shift, date)
            ranked.append(
                {
                    "staff_id": sid,
                    "staff_name": s.name,
                    "contract_type": s.contract_type,
                    "score": float(self._score(sid)),
                    "blocked_by": blocked,
                }
            )
        ranked.sort(key=lambda x: x["score"])
        return ranked[:top_n]

    def _debug_print_no_candidate(self, date: str, shift: ShiftTemplate, role: str) -> None:
        if not self._debug_enabled():
            return
        ranked = []
        for sid, s in self.staff.items():
            if s.role != role:
                continue
            blocked = self._blocked_reasons(sid, role, shift, date)
            ranked.append((s.name, float(self._score(sid)), blocked))
        ranked.sort(key=lambda t: t[1])

        print(f"\n[DEBUG] No chosen candidate for: {date} | {shift.name} | {role}")
        for name, score, blocked in ranked:
            print(f"  - {name} | score={score:.1f} | blocked: {blocked}")

    # ---------- apply assignment ----------
    def _apply_assignment(
        self,
        sid: str,
        role: str,
        shift: ShiftTemplate,
        date: str,
        *,
        allow_double_shift_override: bool = False,
        allow_consecutive_override: bool = False,
    ) -> Tuple[bool, List[str]]:
        reasons = self._blocked_reasons(
            sid, role, shift, date,
            ignore_one_shift_per_day=allow_double_shift_override,
            ignore_consecutive_rules=allow_consecutive_override,
        )

        if any("NON-OVERRIDABLE" in r for r in reasons):
            return False, reasons

        warnings = reasons[:]

        # commit assignment
        a = Assignment(date=date, shift_id=shift.id, role=role, staff_id=sid)
        self.assignments.append(a)
        self._assignments_by_date.setdefault(date, []).append(a)

        self.staff_working_on_date.setdefault(date, set()).add(sid)

        # workload
        hours = self._shift_hours(shift)
        self.staff_hours[sid] += hours
        if shift.is_night:
            self.staff_nights[sid] += 1
        if self._is_weekend(date):
            self.staff_weekends[sid] += 1

        # streak tracking (only for auto logic; projected checks rely on assignment scan)
        prev = self.last_date[sid]
        is_next = prev is not None and self._is_next_day(prev, date)

        if shift.is_night:
            if is_next and self.last_shift_was_night[sid]:
                self.consec_nights[sid] += 1
            else:
                self.consec_nights[sid] = 1
            self.consec_long[sid] = 0
        else:
            if is_next and self.last_shift_was_long[sid]:
                self.consec_long[sid] += 1
            else:
                self.consec_long[sid] = 1
            self.consec_nights[sid] = 0

        self.last_date[sid] = date
        self.last_shift_was_night[sid] = shift.is_night
        self.last_shift_was_long[sid] = (not shift.is_night)

        return True, warnings

    # ---------- selection ----------
    def _pick_best_candidate(self, role: str, shift: ShiftTemplate, date: str) -> Optional[str]:
        candidates: List[str] = []
        for sid, s in self.staff.items():
            if s.role != role:
                continue
            blocked = self._blocked_reasons(sid, role, shift, date)
            if not blocked:
                candidates.append(sid)

        if not candidates:
            return None

        candidates.sort(key=self._score)
        return candidates[0]

    # ---------- generation ----------
    def generate(self) -> None:
        self.assignments.clear()
        self.unfilled.clear()
        self.override_events.clear()
        self.explainability_events.clear()
        self.staff_working_on_date.clear()
        self._assignments_by_date.clear()

        for sid in self.staff.keys():
            self.staff_hours[sid] = 0.0
            self.staff_nights[sid] = 0
            self.staff_weekends[sid] = 0
            self.consec_nights[sid] = 0
            self.consec_long[sid] = 0
            self.last_date[sid] = None
            self.last_shift_was_night[sid] = False
            self.last_shift_was_long[sid] = False

        reqs = sorted(self.requirements, key=lambda r: (r.date, r.shift_id, r.role))
        top_n_explain = int(self.cfg.get("explainability", {}).get("top_n_per_decision", 5))

        for req in reqs:
            shift = self.shift_templates[req.shift_id]
            for _ in range(req.required):
                chosen = self._pick_best_candidate(req.role, shift, req.date)

                if chosen is None:
                    self._debug_print_no_candidate(req.date, shift, req.role)
                    self.unfilled.append(UnfilledSlot(
                        date=req.date,
                        shift_id=req.shift_id,
                        role=req.role,
                        reason="No valid candidate under constraints"
                    ))
                    continue

                ranked = self._rank_candidates_for_decision(req.role, shift, req.date, top_n=top_n_explain)
                ok, _warnings = self._apply_assignment(chosen, req.role, shift, req.date)

                if ok:
                    self._log_explainability(
                        date=req.date,
                        shift=shift,
                        role=req.role,
                        chosen_staff_id=chosen,
                        decision_type="auto",
                        ranked_candidates=ranked,
                        notes="Auto assignment (best valid candidate)",
                    )
                else:
                    self.unfilled.append(UnfilledSlot(
                        date=req.date,
                        shift_id=req.shift_id,
                        role=req.role,
                        reason="Candidate blocked (unexpected) under constraints"
                    ))

        ov = self.cfg.get("overrides", {})
        if ov.get("allow_overrides", True) and ov.get("interactive_overrides", True) and self.unfilled:
            self._interactive_override_unfilled()

    # ---------- cover suggestions ----------
    def _suggest_for_slot(self, slot: UnfilledSlot, top_n: int) -> List[CoverCandidate]:
        shift = self.shift_templates[slot.shift_id]
        candidates: List[CoverCandidate] = []
        bank_only = self._bank_only_consecutive_override_enabled()

        for sid, s in self.staff.items():
            if s.role != slot.role:
                continue

            blocked = self._blocked_reasons(sid, slot.role, shift, slot.date)

            # HARD blocks never allowed
            if any("NON-OVERRIDABLE" in r for r in blocked):
                continue

            # If bank-only mode is enabled, don't suggest permanent staff if the ONLY issue is consecutive exceed
            if bank_only and blocked:
                only_consecutive = all("Would exceed consecutive" in r for r in blocked)
                if only_consecutive and s.contract_type != "bank":
                    continue

            candidates.append(CoverCandidate(staff_id=sid, score=self._score(sid), blocked_by=blocked))

        candidates.sort(key=lambda c: c.score)
        return candidates[:top_n]

    # ---------- interactive helpers ----------
    @staticmethod
    def _prompt_ack(phrase: str) -> bool:
        ans = input(f"Type '{phrase}' to confirm: ").strip()
        return ans.casefold() == phrase.strip().casefold()

    # ---------- reporting ----------
    def print_daily_rota(self) -> None:
        print("\n================ DAILY ROTA ================")
        if not self._assignments_by_date:
            print("No assignments.")
            return
        for date in sorted(self._assignments_by_date.keys()):
            print(f"\n{date}")
            day = sorted(self._assignments_by_date[date], key=lambda a: (a.shift_id, a.role, a.staff_id))
            for a in day:
                s = self.staff[a.staff_id]
                sh = self.shift_templates[a.shift_id]
                print(f"  {sh.name:8} | {a.role:6} | {s.name} ({s.contract_type})")

    def fairness_report(self) -> None:
        print("\n================ FAIRNESS REPORT ================")
        avg_hours = sum(self.staff_hours.values()) / max(1, len(self.staff_hours))
        for sid, s in sorted(self.staff.items(), key=lambda kv: kv[1].name.lower()):
            hours = self.staff_hours[sid]
            nights = self.staff_nights[sid]
            weekends = self.staff_weekends[sid]

            fairness = 100
            if hours > avg_hours + 12:
                fairness -= 20
            if nights > 4:
                fairness -= 20
            if weekends > 3:
                fairness -= 20
            if fairness < 0:
                fairness = 0

            print(f"{s.name:10} | Hours: {hours:5.1f} | Nights: {nights:2d} | Weekends: {weekends:2d} | Fairness: {fairness}")

        c = self.cfg["constraints"]
        print(f"\nUnfilled slots: {len(self.unfilled)}")
        print(f"Max consecutive nights: {c['max_consecutive_nights']}")
        print(f"Max consecutive long days: {c['max_consecutive_long_days']}")

        total_events = len(self.override_events)
        true_overrides = sum(1 for e in self.override_events if e.is_override)
        manual_covers = sum(1 for e in self.override_events if e.is_manual_cover)
        print(f"Logged events: {total_events} | Manual covers: {manual_covers} | True overrides: {true_overrides}")
        print(f"Explainability records: {len(self.explainability_events)}")

    # ---------- interactive override mode ----------
    def _interactive_override_unfilled(self) -> None:
        ov = self.cfg.get("overrides", {})
        warnings_enabled = bool(ov.get("warnings_enabled", True))
        require_ack = bool(ov.get("require_warning_ack", False))
        ack_phrase = str(ov.get("warning_ack_phrase", "I UNDERSTAND"))
        top_n_cover = int(self.cfg["cover_suggestions"]["top_n"])
        top_n_explain = int(self.cfg.get("explainability", {}).get("top_n_per_decision", 5))

        allow_double = bool(ov.get("allow_double_shift", False))
        allow_consec_nights = bool(ov.get("allow_exceed_consecutive_nights", False))
        allow_consec_long = bool(ov.get("allow_exceed_consecutive_long_days", False))
        bank_only = self._bank_only_consecutive_override_enabled()

        print("\n================ OVERRIDE MODE ================")
        print("You can try to cover unfilled slots by overriding SOFT warnings.\n")

        remaining: List[UnfilledSlot] = []

        for slot in self.unfilled:
            shift = self.shift_templates[slot.shift_id]

            print("\n----------------------------------------------")
            print(f"UNFILLED -> Date: {slot.date} | Shift: {shift.name} | Role: {slot.role}")
            if warnings_enabled:
                print(f"Reason: {slot.reason}")

            suggestions = self._suggest_for_slot(slot, top_n=top_n_cover)
            if not suggestions:
                print("No candidates (even with overrides).")
                remaining.append(slot)
                continue

            choices: List[str] = []
            for c in suggestions:
                s = self.staff[c.staff_id]
                blocked_txt = "; ".join(c.blocked_by) if c.blocked_by else "No warnings"
                choices.append(f"{s.name} ({s.contract_type}) | score={c.score:.1f} | {blocked_txt}")

            idx = ask_choice("Pick a candidate to override-assign:", choices, allow_skip=True)
            if idx is None:
                remaining.append(slot)
                continue

            chosen = suggestions[idx]
            staff = self.staff[chosen.staff_id]

            # detect what the choice needs
            needs_double = any("Already working that date" in r for r in chosen.blocked_by)
            needs_consec_raw = any("Would exceed consecutive" in r for r in chosen.blocked_by)

            # bank-only consecutive policy
            needs_consec = bool(needs_consec_raw and (not bank_only or staff.contract_type == "bank"))

            if needs_consec_raw and bank_only and staff.contract_type != "bank":
                print("❌ Consecutive override is restricted to BANK staff by policy.")
                remaining.append(slot)
                continue

            # enforce policy switches from config
            if needs_double and not allow_double:
                print("❌ Not allowed by config: double shifts are disabled (allow_double_shift=false).")
                remaining.append(slot)
                continue

            if needs_consec:
                if shift.is_night and not allow_consec_nights:
                    print("❌ Not allowed by config: exceeding consecutive nights is disabled (allow_exceed_consecutive_nights=false).")
                    remaining.append(slot)
                    continue
                if (not shift.is_night) and not allow_consec_long:
                    print("❌ Not allowed by config: exceeding consecutive long days is disabled (allow_exceed_consecutive_long_days=false).")
                    remaining.append(slot)
                    continue

            did_override = bool(chosen.blocked_by)

            if warnings_enabled and chosen.blocked_by:
                ok = ask_yes_no(f"Override warnings for {staff.name}? ({'; '.join(chosen.blocked_by)})")
                if not ok:
                    remaining.append(slot)
                    continue

                if require_ack and not self._prompt_ack(ack_phrase):
                    print("Override cancelled (phrase mismatch).")
                    remaining.append(slot)
                    continue

            ranked = self._rank_candidates_for_decision(slot.role, shift, slot.date, top_n=top_n_explain)

            ok, warnings = self._apply_assignment(
                sid=chosen.staff_id,
                role=slot.role,
                shift=shift,
                date=slot.date,
                allow_double_shift_override=needs_double and allow_double,
                allow_consecutive_override=needs_consec and (
                    (shift.is_night and allow_consec_nights) or ((not shift.is_night) and allow_consec_long)
                ),
            )

            if ok:
                print(f"✅ Covered -> {slot.date} {shift.name} {slot.role}: {staff.name}")

                self._log_explainability(
                    date=slot.date,
                    shift=shift,
                    role=slot.role,
                    chosen_staff_id=chosen.staff_id,
                    decision_type="manual",
                    ranked_candidates=ranked,
                    notes="Manual cover selection in override mode",
                )

                if did_override:
                    self._log_override(
                        date=slot.date,
                        shift=shift,
                        role=slot.role,
                        staff_id=chosen.staff_id,
                        is_override=True,
                        is_manual_cover=True,
                        override_double=needs_double,
                        override_consecutive=needs_consec,
                        reason_text="; ".join(chosen.blocked_by),
                    )
                else:
                    self._log_override(
                        date=slot.date,
                        shift=shift,
                        role=slot.role,
                        staff_id=chosen.staff_id,
                        is_override=False,
                        is_manual_cover=True,
                        override_double=False,
                        override_consecutive=False,
                        reason_text="Manual cover (no rules broken)",
                    )
            else:
                print(f"❌ Could not assign {staff.name}. Blocks:", "; ".join(warnings))
                remaining.append(slot)

        self.unfilled = remaining

    # =========================
    # AUDIT + QUALITY GATE
    # =========================
    def build_audit(self) -> List[AuditRow]:
        # Prepare per-staff date lists and shift type lists
        by_staff: Dict[str, List[Assignment]] = {sid: [] for sid in self.staff.keys()}
        for a in self.assignments:
            by_staff[a.staff_id].append(a)
        for sid in by_staff:
            by_staff[sid].sort(key=lambda x: (x.date, x.shift_id))

        c = self.cfg["constraints"]
        max_nights_allowed = int(c["max_consecutive_nights"])
        max_long_allowed = int(c["max_consecutive_long_days"])

        avg_hours = sum(self.staff_hours.values()) / max(1, len(self.staff_hours))

        rows: List[AuditRow] = []

        for sid, s in self.staff.items():
            staff_assigns = by_staff.get(sid, [])
            dates = [a.date for a in staff_assigns]
            unique_dates = sorted(set(dates))

            # double shift days (count dates where staff has >1 assignment)
            ds_days = 0
            for d in unique_dates:
                if sum(1 for a in staff_assigns if a.date == d) > 1:
                    ds_days += 1

            # actual max consecutive streaks (scan by date using assignment types)
            def worked_on(d: str, want_night: bool) -> bool:
                for a in self._assignments_by_date.get(d, []):
                    if a.staff_id != sid:
                        continue
                    sh = self.shift_templates[a.shift_id]
                    if sh.is_night == want_night:
                        return True
                return False

            max_night_streak = 0
            max_long_streak = 0

            # scan all dates present in rota range
            all_dates = sorted(self._assignments_by_date.keys())
            # compute streaks
            cur_n = 0
            cur_l = 0
            for d in all_dates:
                if worked_on(d, True):
                    cur_n += 1
                else:
                    cur_n = 0
                if worked_on(d, False):
                    cur_l += 1
                else:
                    cur_l = 0
                max_night_streak = max(max_night_streak, cur_n)
                max_long_streak = max(max_long_streak, cur_l)

            flags: List[str] = []
            if self.staff_hours[sid] > avg_hours + 12:
                flags.append("High hours vs avg")
            if self.staff_nights[sid] > 4:
                flags.append("Many nights")
            if ds_days > 0:
                flags.append("Has double-shift days")
            if max_night_streak > max_nights_allowed:
                flags.append("Exceeded consecutive nights (actual)")
            if max_long_streak > max_long_allowed:
                flags.append("Exceeded consecutive long days (actual)")
            if not flags:
                flags.append("OK")

            rows.append(
                AuditRow(
                    staff_id=sid,
                    staff_name=s.name,
                    role=s.role,
                    contract_type=s.contract_type,
                    target_hours_per_week=s.target_hours_per_week,
                    hours=float(f"{self.staff_hours[sid]:.1f}"),
                    nights=self.staff_nights[sid],
                    weekends=self.staff_weekends[sid],
                    shifts_total=len(staff_assigns),
                    days_worked=len(unique_dates),
                    double_shift_days=ds_days,
                    max_consecutive_nights_actual=max_night_streak,
                    max_consecutive_long_days_actual=max_long_streak,
                    overload_flags="; ".join(flags),
                )
            )

        # stable output ordering
        rows.sort(key=lambda r: r.staff_name.lower())
        return rows

    def quality_gate(self, audit_rows: List[AuditRow]) -> Tuple[bool, List[str]]:
        issues: List[str] = []
        ov = self.cfg.get("overrides", {})
        c = self.cfg.get("constraints")

        if len(self.unfilled) > 0:
            issues.append(f"Unfilled slots remain: {len(self.unfilled)}")

        allow_double = bool(ov.get("allow_double_shift", False))
        allow_consec_nights = bool(ov.get("allow_exceed_consecutive_nights", False))
        allow_consec_long = bool(ov.get("allow_exceed_consecutive_long_days", False))
        bank_only = self._bank_only_consecutive_override_enabled()

        one_shift_rule = bool(c.get("one_shift_per_day", True))
        max_n = int(c["max_consecutive_nights"])
        max_l = int(c["max_consecutive_long_days"])

        # double shift policy
        if one_shift_rule and not allow_double:
            offenders = [r.staff_name for r in audit_rows if r.double_shift_days > 0]
            if offenders:
                issues.append(f"Double-shift days exist but policy forbids them: {', '.join(offenders)}")

        # consecutive policy
        for r in audit_rows:
            # nights
            if r.max_consecutive_nights_actual > max_n:
                if bank_only and r.contract_type != "bank":
                    issues.append(f"{r.staff_name} exceeded consecutive nights, but bank-only policy forbids permanent staff.")
                elif not allow_consec_nights:
                    issues.append(f"{r.staff_name} exceeded consecutive nights, but policy forbids it.")
            # long
            if r.max_consecutive_long_days_actual > max_l:
                if bank_only and r.contract_type != "bank":
                    issues.append(f"{r.staff_name} exceeded consecutive long days, but bank-only policy forbids permanent staff.")
                elif not allow_consec_long:
                    issues.append(f"{r.staff_name} exceeded consecutive long days, but policy forbids it.")

        return (len(issues) == 0), issues

    # =========================
    # EXPORTS
    # =========================
    def export_files(self) -> None:
        ex = self.cfg["exports"]
        folder = ex["export_folder"]
        os.makedirs(folder, exist_ok=True)

        exported_csv: List[str] = []
        exported_excel: Optional[str] = None

        audit_rows = self.build_audit()
        passed, issues = self.quality_gate(audit_rows)

        if ex.get("export_csv", True):
            exported_csv.extend(self._export_csv(folder))
            if ex.get("export_overrides_csv", True):
                exported_csv.append(self._export_overrides_csv(folder))
            if ex.get("export_explainability_csv", True):
                exported_csv.append(self._export_explainability_csv(folder))
            if ex.get("export_audit_csv", True):
                exported_csv.append(self._export_audit_csv(folder, audit_rows))

        if ex.get("export_excel", True):
            exported_excel = self._export_excel(folder, audit_rows)

        print("\n================ QUALITY GATE ================")
        if passed:
            print("✅ PASS: rota meets current policy settings.")
        else:
            print("❌ FAIL: rota violates current policy settings:")
            for i, msg in enumerate(issues, 1):
                print(f"  {i}) {msg}")

        print(f"Audit rows: {len(audit_rows)}")

        if exported_csv:
            print("✅ CSV files exported:")
            for f in exported_csv:
                print(f" - {f}")

        if exported_excel:
            print("✅ Excel exported:")
            print(f" - {exported_excel}")

    def _export_csv(self, folder: str) -> List[str]:
        out: List[str] = []

        a_path = os.path.join(folder, "rota_assignments.csv")
        with open(a_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["date", "shift_id", "shift_name", "role", "staff_id", "staff_name", "contract_type"])
            for a in sorted(self.assignments, key=lambda x: (x.date, x.shift_id, x.role, x.staff_id)):
                s = self.staff[a.staff_id]
                sh = self.shift_templates[a.shift_id]
                w.writerow([a.date, a.shift_id, sh.name, a.role, a.staff_id, s.name, s.contract_type])
        out.append("rota_assignments.csv")

        u_path = os.path.join(folder, "rota_unfilled.csv")
        with open(u_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["date", "shift_id", "shift_name", "role", "reason"])
            for u in self.unfilled:
                sh = self.shift_templates[u.shift_id]
                w.writerow([u.date, u.shift_id, sh.name, u.role, u.reason])
        out.append("rota_unfilled.csv")

        fr_path = os.path.join(folder, "rota_fairness.csv")
        with open(fr_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["staff_id", "name", "role", "contract_type", "hours", "nights", "weekends"])
            for sid, s in sorted(self.staff.items(), key=lambda kv: kv[1].name.lower()):
                w.writerow([sid, s.name, s.role, s.contract_type, f"{self.staff_hours[sid]:.1f}", self.staff_nights[sid], self.staff_weekends[sid]])
        out.append("rota_fairness.csv")

        return out

    def _export_overrides_csv(self, folder: str) -> str:
        path = os.path.join(folder, "rota_overrides.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "timestamp_utc",
                "date",
                "shift_id",
                "shift_name",
                "role",
                "staff_id",
                "staff_name",
                "contract_type",
                "is_override",
                "is_manual_cover",
                "override_double_shift",
                "override_consecutive",
                "override_reason_text",
            ])
            for e in self.override_events:
                w.writerow([
                    e.timestamp_utc,
                    e.date,
                    e.shift_id,
                    e.shift_name,
                    e.role,
                    e.staff_id,
                    e.staff_name,
                    e.contract_type,
                    e.is_override,
                    e.is_manual_cover,
                    e.override_double_shift,
                    e.override_consecutive,
                    e.override_reason_text,
                ])
        return "rota_overrides.csv"

    def _export_explainability_csv(self, folder: str) -> str:
        path = os.path.join(folder, "rota_explainability.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "timestamp_utc",
                "date",
                "shift_id",
                "shift_name",
                "role",
                "chosen_staff_id",
                "chosen_staff_name",
                "chosen_score",
                "decision_type",
                "top_n",
                "ranked_candidates_json",
                "notes",
            ])
            for e in self.explainability_events:
                w.writerow([
                    e.timestamp_utc,
                    e.date,
                    e.shift_id,
                    e.shift_name,
                    e.role,
                    e.chosen_staff_id,
                    e.chosen_staff_name,
                    e.chosen_score,
                    e.decision_type,
                    e.top_n,
                    e.ranked_candidates_json,
                    e.notes,
                ])
        return "rota_explainability.csv"

    def _export_audit_csv(self, folder: str, audit_rows: List[AuditRow]) -> str:
        path = os.path.join(folder, "rota_audit.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([
                "staff_id",
                "staff_name",
                "role",
                "contract_type",
                "target_hours_per_week",
                "hours",
                "nights",
                "weekends",
                "shifts_total",
                "days_worked",
                "double_shift_days",
                "max_consecutive_nights_actual",
                "max_consecutive_long_days_actual",
                "overload_flags",
            ])
            for r in audit_rows:
                w.writerow([
                    r.staff_id,
                    r.staff_name,
                    r.role,
                    r.contract_type,
                    r.target_hours_per_week,
                    f"{r.hours:.1f}",
                    r.nights,
                    r.weekends,
                    r.shifts_total,
                    r.days_worked,
                    r.double_shift_days,
                    r.max_consecutive_nights_actual,
                    r.max_consecutive_long_days_actual,
                    r.overload_flags,
                ])
        return "rota_audit.csv"

    def _export_excel(self, folder: str, audit_rows: List[AuditRow]) -> Optional[str]:
        try:
            from openpyxl import Workbook
        except Exception:
            print("⚠️ Excel export skipped: openpyxl not installed (run: pip install openpyxl)")
            return None

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Assignments"
        ws1.append(["date", "shift_id", "shift_name", "role", "staff_id", "staff_name", "contract_type"])
        for a in sorted(self.assignments, key=lambda x: (x.date, x.shift_id, x.role, x.staff_id)):
            s = self.staff[a.staff_id]
            sh = self.shift_templates[a.shift_id]
            ws1.append([a.date, a.shift_id, sh.name, a.role, a.staff_id, s.name, s.contract_type])

        ws2 = wb.create_sheet("Unfilled")
        ws2.append(["date", "shift_id", "shift_name", "role", "reason"])
        for u in self.unfilled:
            sh = self.shift_templates[u.shift_id]
            ws2.append([u.date, u.shift_id, sh.name, u.role, u.reason])

        ws3 = wb.create_sheet("Fairness")
        ws3.append(["staff_id", "name", "role", "contract_type", "hours", "nights", "weekends"])
        for sid, s in sorted(self.staff.items(), key=lambda kv: kv[1].name.lower()):
            ws3.append([sid, s.name, s.role, s.contract_type, float(f"{self.staff_hours[sid]:.1f}"), self.staff_nights[sid], self.staff_weekends[sid]])

        ws4 = wb.create_sheet("Overrides")
        ws4.append([
            "timestamp_utc",
            "date",
            "shift_id",
            "shift_name",
            "role",
            "staff_id",
            "staff_name",
            "contract_type",
            "is_override",
            "is_manual_cover",
            "override_double_shift",
            "override_consecutive",
            "override_reason_text",
        ])
        for e in self.override_events:
            ws4.append([
                e.timestamp_utc,
                e.date,
                e.shift_id,
                e.shift_name,
                e.role,
                e.staff_id,
                e.staff_name,
                e.contract_type,
                e.is_override,
                e.is_manual_cover,
                e.override_double_shift,
                e.override_consecutive,
                e.override_reason_text,
            ])

        ws5 = wb.create_sheet("Explainability")
        ws5.append([
            "timestamp_utc",
            "date",
            "shift_id",
            "shift_name",
            "role",
            "chosen_staff_id",
            "chosen_staff_name",
            "chosen_score",
            "decision_type",
            "top_n",
            "ranked_candidates_json",
            "notes",
        ])
        for e in self.explainability_events:
            ws5.append([
                e.timestamp_utc,
                e.date,
                e.shift_id,
                e.shift_name,
                e.role,
                e.chosen_staff_id,
                e.chosen_staff_name,
                e.chosen_score,
                e.decision_type,
                e.top_n,
                e.ranked_candidates_json,
                e.notes,
            ])

        ws6 = wb.create_sheet("Audit")
        ws6.append([
            "staff_id",
            "staff_name",
            "role",
            "contract_type",
            "target_hours_per_week",
            "hours",
            "nights",
            "weekends",
            "shifts_total",
            "days_worked",
            "double_shift_days",
            "max_consecutive_nights_actual",
            "max_consecutive_long_days_actual",
            "overload_flags",
        ])
        for r in audit_rows:
            ws6.append([
                r.staff_id,
                r.staff_name,
                r.role,
                r.contract_type,
                r.target_hours_per_week,
                float(f"{r.hours:.1f}"),
                r.nights,
                r.weekends,
                r.shifts_total,
                r.days_worked,
                r.double_shift_days,
                r.max_consecutive_nights_actual,
                r.max_consecutive_long_days_actual,
                r.overload_flags,
            ])

        out_path = os.path.join(folder, "rota.xlsx")
        wb.save(out_path)
        return "rota.xlsx"


# =========================
# BUILD INPUTS (CSV preferred, demo fallback)
# =========================
def build_inputs() -> Tuple[Dict[str, ShiftTemplate], List[Staff], List[Requirement]]:
    long_day = ShiftTemplate("long", "Long Day", "08:00", "20:00", False)
    night = ShiftTemplate("night", "Night", "20:00", "08:00", True)
    shifts = {"long": long_day, "night": night}

    if os.path.exists("staff.csv") and os.path.exists("requirements.csv"):
        staff = load_staff_csv("staff.csv")
        reqs = load_requirements_csv("requirements.csv")
        return shifts, staff, reqs

    # demo fallback (same as your earlier demo)
    staff = [
        Staff("s1", "Amaka", "HCA", "permanent", 44, True),
        Staff("s2", "James", "HCA", "permanent", 44, True),
        Staff("s3", "Fatima", "HCA", "permanent", 33, False),
        Staff("s4", "Daniel", "HCA", "permanent", 44, True),
        Staff("s5", "Grace", "HCA", "bank", 0, False),
        Staff("s6", "Uche", "HCA", "bank", 0, True),
        Staff("s7", "Ife", "Senior", "permanent", 33, True),
        Staff("s8", "Tunde", "Senior", "permanent", 44, True),
        Staff("s9", "Helen", "Senior", "bank", 0, True),
        Staff("s10", "Zara", "Senior", "bank", 0, False),
    ]

    requirements: List[Requirement] = []
    start = datetime.strptime("2025-01-06", DATE_FMT)
    for i in range(7):
        d = (start + timedelta(days=i)).strftime(DATE_FMT)
        requirements += [
            Requirement(d, "long", "Senior", 1),
            Requirement(d, "long", "HCA", 3),
            Requirement(d, "night", "Senior", 1),
            Requirement(d, "night", "HCA", 2),
        ]

    return shifts, staff, requirements


# =========================
# MAIN
# =========================
def demo() -> None:
    cfg = load_config("config.json")
    print("Running demo rota...")
    print("Config loaded from config.json")

    shifts, staff, requirements = build_inputs()

    gen = RotaGenerator(cfg, shifts, staff, requirements)
    gen.generate()

    print("\nTotal assignments:", len(gen.assignments))
    for a in gen.assignments[:10]:
        print(a)

    gen.print_daily_rota()
    gen.fairness_report()
    gen.export_files()


if __name__ == "__main__":
    demo()
